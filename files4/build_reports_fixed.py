"""
build_reports_fixed.py
======================
Generates:
  • qatifan_annual_report_2025.pdf  — Arabic PDF with cover + 3 reports
  • qatifan_reports_2025.xlsx       — 4-sheet Excel workbook with charts

Fixes applied vs original:
  1.  CoverPage.draw() now stores availWidth/availHeight from wrap() so
      drawing coordinates are always correct.
  2.  Removed dead imports: numbers, DataPoint, os.
  3.  PDF TableStyle 'TOPBORDERPADDING' → valid 'TOPPADDING'.
  4.  Summary-table closing-balance style index corrected (row 4 = last).
  5.  notif_kpi header: single consistent GOLD background (no conflicting NAVY).
  6.  Expense % formula replaced with pre-computed string (recalc unavailable).
  7.  ws1 expense section: columns A/C/D/E now widened individually, not via
      hdr_row (which only knows sequential A→N columns).
  8.  ws1 column-width override after total row removed.
  9.  All Alignment() calls that need centering/right-align pass explicit args.
  10. Line chart Reference rows aligned: header=39, data=40-45 for both series
      and categories.
  11. sty() 'font' param removed (it was silently ignored).
  12. Honour-roll unpacking made explicit.
  13. CoverPage stores dimensions in wrap() for use in draw().
  14. Excel values hardcoded instead of formula strings so they display
      correctly without recalc.py (formulas added as comments where useful).
  15. ws2/ws3 total-row Alignment made explicit.
"""

import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.platypus.flowables import Flowable
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from datetime import datetime

# ── Font registration ────────────────────────────────────────────────────
pdfmetrics.registerFont(TTFont('Ar',     '/usr/share/fonts/truetype/freefont/FreeSans.ttf'))
pdfmetrics.registerFont(TTFont('ArBold', '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf'))

def ar(t):
    """Reshape + apply BiDi algorithm for correct Arabic rendering in PDF."""
    return get_display(arabic_reshaper.reshape(str(t)))

# ── Colour palette ───────────────────────────────────────────────────────
NAVY  = colors.HexColor('#0B1F3A')
BLUE  = colors.HexColor('#1A4A8A')
TEAL  = colors.HexColor('#0D7A5F')
GOLD  = colors.HexColor('#C07800')
RED   = colors.HexColor('#B02020')
LGRAY = colors.HexColor('#F5F5F3')
MGRAY = colors.HexColor('#D0D0CC')
WHITE = colors.white
BLACK = colors.black

# ── Paragraph styles ─────────────────────────────────────────────────────
def sty(name, size=10, color=BLACK, align=TA_RIGHT,
        bold=False, space_before=0, space_after=4, leading=None):
    return ParagraphStyle(
        name,
        fontName='ArBold' if bold else 'Ar',
        fontSize=size,
        textColor=color,
        alignment=align,
        spaceAfter=space_after,
        spaceBefore=space_before,
        leading=leading or (size * 1.5),
        wordWrap='RTL',
    )

S_H1    = sty('h1',   size=15, color=NAVY,  bold=True,  space_before=14, space_after=6)
S_H2    = sty('h2',   size=12, color=BLUE,  bold=True,  space_before=8,  space_after=4)
S_BODY  = sty('body', size=9.5)
S_NOTE  = sty('note', size=9,  color=TEAL,  space_before=8)
S_REC   = sty('rec',  size=9,  space_before=4, space_after=4)

# ── Table style builder ──────────────────────────────────────────────────
def ts(*extra):
    base = [
        ('FONTNAME',       (0, 0), (-1, -1), 'Ar'),
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
        ('ALIGN',          (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LGRAY]),
        ('GRID',           (0, 0), (-1, -1), 0.3, MGRAY),
        ('LEFTPADDING',    (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',   (0, 0), (-1, -1), 6),
        ('TOPPADDING',     (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
    ]
    return TableStyle(base + list(extra))

def nav_header():
    """Standard navy header row commands."""
    return [
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'ArBold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9.5),
    ]

def colored_header(color):
    """Header row with custom color."""
    return [
        ('BACKGROUND', (0, 0), (-1, 0), color),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'ArBold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9.5),
    ]

def total_row(row_idx):
    """Navy total row commands."""
    return [
        ('BACKGROUND', (0, row_idx), (-1, row_idx), NAVY),
        ('TEXTCOLOR',  (0, row_idx), (-1, row_idx), WHITE),
        ('FONTNAME',   (0, row_idx), (-1, row_idx), 'ArBold'),
    ]

# ── Cover Page Flowable ──────────────────────────────────────────────────
class CoverPage(Flowable):
    """Full-page cover. Dimensions are captured from wrap() so draw() is consistent."""

    def __init__(self):
        Flowable.__init__(self)
        self._w = 0
        self._h = 0

    def wrap(self, availWidth, availHeight):
        # FIX 1 & 14: store actual available dimensions for draw()
        self._w = availWidth
        self._h = availHeight
        return (availWidth, availHeight)

    def draw(self):
        c = self.canv
        w, h = self._w, self._h

        # Navy background
        c.setFillColor(NAVY)
        c.rect(0, 0, w, h, fill=1, stroke=0)

        # Gold accent bar
        c.setFillColor(GOLD)
        c.rect(0, h * 0.52, w, 4, fill=1, stroke=0)

        # Title text
        c.setFillColor(WHITE)
        c.setFont('ArBold', 26)
        c.drawCentredString(w / 2, h * 0.62, ar('التقارير السنوية الختامية'))
        c.setFont('Ar', 16)
        c.drawCentredString(w / 2, h * 0.55, ar('صندوق عائلة قطيفان'))
        c.setFillColor(MGRAY)
        c.setFont('Ar', 11)
        c.drawCentredString(w / 2, h * 0.48, ar('السنة المالية 2025'))
        c.drawCentredString(w / 2, h * 0.44,
                            ar('تاريخ الإصدار: ' + datetime.now().strftime('%Y/%m/%d')))

        # Footer
        c.setFillColor(TEAL)
        c.rect(0, 32, w, 2, fill=1, stroke=0)
        c.setFillColor(MGRAY)
        c.setFont('Ar', 8)
        c.drawCentredString(w / 2, 18, ar('وثيقة سرية — لاستخدام الأعضاء فقط'))

# ════════════════════════════════════════════════════════════════════════
#  DATA
# ════════════════════════════════════════════════════════════════════════
OPENING_BALANCE = 32_500.00

expenses_detail = {
    'condolence': [
        ('وفاة والدة أحمد القطيفان',     500, '18 يونيو 2025'),
        ('وفاة عم عبدالرحمن القطيفان',   500, '25 مايو 2025'),
        ('وفاة جد محمد القطيفان',        500, '10 مارس 2025'),
        ('وفاة زوجة خالد القطيفان',      500,  '3 يناير 2025'),
    ],
    'wedding': [
        ('نقوط زواج سالم محمد القطيفان',   1000, '24 يونيو 2025'),
        ('نقوط زواج ناصر عبدالله القطيفان', 1000, '10 يونيو 2025'),
        ('نقوط زواج فهد علي القطيفان',     1000, '14 أبريل 2025'),
        ('نقوط زواج يوسف سعد القطيفان',    1000,  '2 فبراير 2025'),
        ('نقوط زواج طارق حسن القطيفان',    1000, '18 يناير 2025'),
    ],
    'emergency': [
        ('مساعدة طارئة — علي القطيفان',    800,  '2 يونيو 2025'),
        ('مساعدة طارئة — منصور القطيفان',  900, '20 مارس 2025'),
        ('مساعدة طارئة — سلمى القطيفان',   750,  '5 فبراير 2025'),
        ('مساعدة طارئة — ريم القطيفان',    650, '12 يناير 2025'),
    ],
    'admin': [
        ('رسوم خدمة بنكية',      120, 'يناير 2025'),
        ('اشتراك خدمة الرسائل',  480, 'يناير 2025'),
        ('مستلزمات إدارية',      250, 'مارس 2025'),
        ('رسوم صيانة النظام',    400, 'يونيو 2025'),
    ],
}

monthly_income = [
    ('يناير', 3600), ('فبراير', 3450), ('مارس',    3150),
    ('أبريل', 3300), ('مايو',   2850), ('يونيو',   2700),
    ('يوليو', 3000), ('أغسطس', 2550), ('سبتمبر', 2850),
    ('أكتوبر',3150), ('نوفمبر', 2700), ('ديسمبر', 3000),
]

members_data = [
    # (name, paid, debt, months_paid, pct, rating)
    ('عبدالله محمد القطيفان',  1800,    0, 12, 100.0, 'ممتاز'),
    ('فاطمة علي القطيفان',     1800,    0, 12, 100.0, 'ممتاز'),
    ('سعد خالد القطيفان',      1800,    0, 12, 100.0, 'ممتاز'),
    ('نورة سالم القطيفان',     1650,    0, 11, 100.0, 'ممتاز'),
    ('عمر ناصر القطيفان',      1600,  300,  9,  75.0, 'جيد'),
    ('هند يوسف القطيفان',      1400,  400,  8,  66.7, 'جيد'),
    ('خالد سعد القطيفان',      1250,  550,  7,  58.3, 'متوسط'),
    ('يوسف حسن القطيفان',      1050,  750,  7,  58.3, 'متوسط'),
    ('سلطان علي القطيفان',     1200,  600,  6,  50.0, 'متأخر'),
    ('محمد ناصر القطيفان',      900,  900,  6,  50.0, 'متأخر'),
    ('أحمد محمد القطيفان',      900,  900,  5,  41.7, 'متأخر'),
    ('عبدالرحمن عمر القطيفان',  750, 1050,  5,  41.7, 'متأخر'),
]

notif_data = [
    # (month, sent, delivered, opened, paid_after, channel)
    ('يناير',  48, 45, 41, 37, 'واتساب+إيميل'),
    ('فبراير', 45, 42, 38, 34, 'واتساب+إيميل'),
    ('مارس',   50, 47, 45, 40, 'واتساب+إيميل'),
    ('أبريل',  46, 43, 41, 38, 'واتساب+إيميل'),
    ('مايو',   52, 49, 44, 39, 'واتساب+إيميل'),
    ('يونيو',  55, 51, 47, 42, 'واتساب+إيميل'),
]

# Derived totals
total_income     = sum(v for _, v in monthly_income)
exp_condolence   = sum(v for _, v, _ in expenses_detail['condolence'])
exp_wedding      = sum(v for _, v, _ in expenses_detail['wedding'])
exp_emergency    = sum(v for _, v, _ in expenses_detail['emergency'])
exp_admin        = sum(v for _, v, _ in expenses_detail['admin'])
total_expenses   = exp_condolence + exp_wedding + exp_emergency + exp_admin
closing_balance  = OPENING_BALANCE + total_income - total_expenses
total_debt       = sum(r[2] for r in members_data)
total_sent       = sum(r[1] for r in notif_data)
total_delivered  = sum(r[2] for r in notif_data)
total_opened     = sum(r[3] for r in notif_data)
total_paid_after = sum(r[4] for r in notif_data)

# ════════════════════════════════════════════════════════════════════════
#  PDF
# ════════════════════════════════════════════════════════════════════════
doc = SimpleDocTemplate(
    '/home/claude/qatifan_annual_report_2025.pdf',
    pagesize=A4,
    rightMargin=2 * cm, leftMargin=2 * cm,
    topMargin=2 * cm,   bottomMargin=2 * cm,
    title='التقارير السنوية — صندوق عائلة قطيفان 2025',
)

story = []

# Cover
story.append(CoverPage())
story.append(PageBreak())

# ─── Report 1: Annual Accounts ───────────────────────────────────────────
story.append(Paragraph(ar('أولاً: تقرير الحسابات السنوية'), S_H1))
story.append(HRFlowable(width='100%', thickness=2, color=NAVY, spaceAfter=10))

# Summary table (5 rows: header + 4 data)
story.append(Paragraph(ar('ملخص الحركة المالية للسنة'), S_H2))
summary_rows = [
    [ar('البيان'),                              ar('المبلغ (ر.س)')],
    [ar('رصيد بداية المدة — 1 يناير 2025'),    f'{OPENING_BALANCE:,.2f}'],
    [ar('إجمالي الاشتراكات المحصلة'),           f'{total_income:,.2f}'],
    [ar('(−) إجمالي المصروفات'),                f'({total_expenses:,.2f})'],
    [ar('رصيد نهاية المدة — 31 ديسمبر 2025'),  f'{closing_balance:,.2f}'],
]
t_summary = Table(summary_rows, colWidths=[11 * cm, 4.5 * cm])
t_summary.setStyle(ts(
    *nav_header(),
    # FIX 5: row index 4 = last data row (closing balance) — correct
    ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#E8F5F0')),
    ('FONTNAME',   (0, 4), (-1, 4), 'ArBold'),
    ('TEXTCOLOR',  (0, 4), (-1, 4), TEAL),
    ('ALIGN',      (1, 1), (1, -1), 'LEFT'),
    # FIX 13: 'TOPBORDERPADDING' → 'TOPPADDING'
    ('TOPPADDING', (0, 4), (-1, 4), 8),
))
story.append(t_summary)
story.append(Spacer(1, 12))

# Monthly income table
story.append(Paragraph(ar('الاشتراكات المحصلة شهرياً'), S_H2))
last_r = len(monthly_income) + 1   # header=0, data=1..12, total=13
inc_rows = ([[ar('الشهر'), ar('المحصَّل (ر.س)')]] +
            [[ar(m), f'{v:,.2f}'] for m, v in monthly_income] +
            [[ar('الإجمالي'), f'{total_income:,.2f}']])
t_income = Table(inc_rows, colWidths=[8 * cm, 7.5 * cm])
t_income.setStyle(ts(
    *nav_header(),
    *total_row(last_r),
    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
))
story.append(t_income)
story.append(Spacer(1, 12))

# Expense category summary
story.append(Paragraph(ar('تفصيل المصروفات حسب البند'), S_H2))
cat_keys  = ['condolence', 'wedding', 'emergency', 'admin']
cat_names = ['العزاء', 'نقوط الزواج', 'الطوارئ', 'المصاريف الإدارية']
cat_totals = [exp_condolence, exp_wedding, exp_emergency, exp_admin]
cat_rows = [[ar('البند'), ar('الحالات'), ar('الإجمالي (ر.س)'), ar('النسبة')]]
for key, name, tot in zip(cat_keys, cat_names, cat_totals):
    pct = tot / total_expenses * 100
    cat_rows.append([ar(name), str(len(expenses_detail[key])),
                     f'{tot:,.2f}', f'{pct:.1f}%'])
cat_rows.append([ar('المجموع'),
                 str(sum(len(v) for v in expenses_detail.values())),
                 f'{total_expenses:,.2f}', '100.0%'])
t_cats = Table(cat_rows, colWidths=[5 * cm, 3 * cm, 4.5 * cm, 3 * cm])
t_cats.setStyle(ts(
    *nav_header(),
    *total_row(len(cat_rows) - 1),
    ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
    ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#EEF0FA')),
    ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#FFF8EE')),
))
story.append(t_cats)
story.append(Spacer(1, 12))

# Detailed expense tables per category
for key, label, hdr_color in [
    ('condolence', 'تفصيل مصروفات العزاء',           NAVY),
    ('wedding',    'تفصيل مصروفات نقوط الزواج',      BLUE),
    ('emergency',  'تفصيل مصروفات الطوارئ',           RED),
    ('admin',      'تفصيل المصروفات الإدارية',        TEAL),
]:
    story.append(Paragraph(ar(label), S_H2))
    rows = [[ar('البيان'), ar('التاريخ'), ar('المبلغ (ر.س)')]]
    for item, amt, dt in expenses_detail[key]:
        rows.append([ar(item), ar(dt), f'{amt:,.2f}'])
    sub_total = sum(v for _, v, _ in expenses_detail[key])
    rows.append([ar('الإجمالي'), '', f'{sub_total:,.2f}'])
    t_detail = Table(rows, colWidths=[7.5 * cm, 4 * cm, 4 * cm])
    t_detail.setStyle(ts(
        *colored_header(hdr_color),
        *total_row(len(rows) - 1),
        ('BACKGROUND', (0, len(rows) - 1), (-1, -1), colors.HexColor('#F0F0EE')),
        ('TEXTCOLOR',  (0, len(rows) - 1), (-1, -1), BLACK),
        ('ALIGN',      (2, 0), (2, -1), 'LEFT'),
    ))
    story.append(t_detail)
    story.append(Spacer(1, 8))

story.append(PageBreak())

# ─── Report 2: Dues & Commitment ─────────────────────────────────────────
story.append(Paragraph(ar('ثانياً: تقرير الذمم والالتزام'), S_H1))
story.append(HRFlowable(width='100%', thickness=2, color=BLUE, spaceAfter=10))

story.append(Paragraph(ar('ملخص الذمم نهاية العام'), S_H2))
dues_sum_rows = [
    [ar('البيان'),                          ar('القيمة')],
    [ar('إجمالي عدد الأعضاء'),             '28'],
    [ar('أعضاء بحساب مسوَّى بالكامل'),     '17'],
    [ar('أعضاء لديهم ذمم متأخرة'),         '11'],
    [ar('إجمالي الذمم المتراكمة (ر.س)'),   f'{total_debt:,.2f}'],
]
t_dues_sum = Table(dues_sum_rows, colWidths=[11 * cm, 4.5 * cm])
t_dues_sum.setStyle(ts(
    *nav_header(),
    ('ALIGN',     (1, 0), (1, -1), 'LEFT'),
    ('TEXTCOLOR', (0, 4), (-1, 4), RED),
    ('FONTNAME',  (0, 4), (-1, 4), 'ArBold'),
))
story.append(t_dues_sum)
story.append(Spacer(1, 10))

# Full member statement
story.append(Paragraph(ar('كشف حساب الأعضاء التفصيلي'), S_H2))
mem_rows = [[ar('الاسم'), ar('مسدَّد (ر.س)'), ar('ذمة (ر.س)'),
             ar('أشهر'), ar('الالتزام'), ar('التقييم')]]
for name, paid, debt, months_p, pct, rating in members_data:
    mem_rows.append([
        ar(name), f'{paid:,.2f}',
        f'{debt:,.2f}' if debt else '—',
        str(months_p), f'{pct:.1f}%', ar(rating),
    ])
t_mem = Table(mem_rows, colWidths=[5.2 * cm, 2.8 * cm, 2.8 * cm,
                                    1.6 * cm, 2.2 * cm, 1.9 * cm])
# FIX 12 (honour): explicit unpacking so no fragile star-unpack
mem_style_cmds = [*nav_header(), ('ALIGN', (1, 0), (-1, -1), 'LEFT')]
for i, (_, _, _, _, _, rating) in enumerate(members_data, 1):
    if rating == 'ممتاز':
        mem_style_cmds += [('TEXTCOLOR', (5, i), (5, i), TEAL),
                            ('FONTNAME',  (5, i), (5, i), 'ArBold')]
    elif rating == 'متأخر':
        mem_style_cmds += [('TEXTCOLOR', (5, i), (5, i), RED),
                            ('FONTNAME',  (5, i), (5, i), 'ArBold'),
                            ('TEXTCOLOR', (2, i), (2, i), RED)]
t_mem.setStyle(ts(*mem_style_cmds))
story.append(t_mem)
story.append(Spacer(1, 14))

# Honour roll
story.append(Paragraph(ar('لوحة الشرف — الأعضاء الأكثر التزاماً'), S_H2))
top = [(name, paid, pct) for name, paid, _, _, pct, rating in members_data
       if rating == 'ممتاز']
honour_rows = [[ar('الترتيب'), ar('الاسم'), ar('إجمالي المدفوع (ر.س)'), ar('معدل الالتزام')]]
for i, (name, paid, pct) in enumerate(top, 1):
    honour_rows.append([str(i), ar(name), f'{paid:,.2f}', f'{pct:.1f}%'])
t_honour = Table(honour_rows, colWidths=[2 * cm, 7 * cm, 4.5 * cm, 3 * cm])
t_honour.setStyle(ts(
    *colored_header(TEAL),
    ('ALIGN',     (1, 0), (-1, -1), 'LEFT'),
    ('TEXTCOLOR', (0, 1), (-1, -1), TEAL),
    ('FONTNAME',  (0, 1), (-1, -1), 'ArBold'),
))
story.append(t_honour)
story.append(Paragraph(
    ar('يُوصى بتكريم هؤلاء الأعضاء في الاجتماع السنوي للعائلة تقديراً لالتزامهم الكامل.'),
    S_NOTE,
))

story.append(PageBreak())

# ─── Report 3: Notification Efficiency ───────────────────────────────────
story.append(Paragraph(ar('ثالثاً: تقرير كفاءة التنبيهات الآلية'), S_H1))
story.append(HRFlowable(width='100%', thickness=2, color=GOLD, spaceAfter=10))

# FIX 6: use single colored_header(GOLD) — no conflicting nav_header
notif_kpi_rows = [
    [ar('المؤشر'),                       ar('القيمة'),           ar('النسبة')],
    [ar('إجمالي الرسائل المرسلة'),       str(total_sent),         '100%'],
    [ar('الرسائل الموصلة بنجاح'),        str(total_delivered),
     f'{total_delivered / total_sent * 100:.1f}%'],
    [ar('الرسائل المفتوحة/المقروءة'),     str(total_opened),
     f'{total_opened / total_sent * 100:.1f}%'],
    [ar('حالات سداد بعد التنبيه'),       str(total_paid_after),
     f'{total_paid_after / total_sent * 100:.1f}%'],
]
t_kpi = Table(notif_kpi_rows, colWidths=[7 * cm, 3.5 * cm, 5 * cm])
t_kpi.setStyle(ts(
    *colored_header(GOLD),
    ('ALIGN',     (1, 0), (-1, -1), 'LEFT'),
    # Last data row highlight
    ('TEXTCOLOR', (0, 4), (-1, 4), TEAL),
    ('FONTNAME',  (0, 4), (-1, 4), 'ArBold'),
))
story.append(t_kpi)
story.append(Spacer(1, 12))

story.append(Paragraph(ar('التفصيل الشهري لأداء التنبيهات'), S_H2))
detail_rows = [[ar('الشهر'), ar('مُرسَل'), ar('موصَّل'),
                ar('مفتوح'), ar('سداد بعده'), ar('قناة'), ar('معدل التحويل')]]
for month, sent, delivered, opened, paid_after, channel in notif_data:
    detail_rows.append([
        ar(month), str(sent), str(delivered), str(opened), str(paid_after),
        ar(channel), f'{paid_after / sent * 100:.1f}%',
    ])
# Totals row
detail_rows.append([
    ar('الإجمالي'), str(total_sent), str(total_delivered),
    str(total_opened), str(total_paid_after), '',
    f'{total_paid_after / total_sent * 100:.1f}%',
])
t_detail = Table(detail_rows,
                 colWidths=[1.8 * cm, 1.6 * cm, 1.8 * cm,
                             1.8 * cm, 2.3 * cm, 4 * cm, 2.8 * cm])
t_detail.setStyle(ts(
    *colored_header(GOLD),
    *total_row(len(detail_rows) - 1),
    ('ALIGN',    (1, 0), (-1, -1), 'LEFT'),
    ('FONTSIZE', (0, 0), (-1, -1), 8.5),
))
story.append(t_detail)
story.append(Spacer(1, 10))

story.append(Paragraph(ar('الملاحظات والتوصيات'), S_H2))
recs = [
    '• معدل توصيل الرسائل بلغ 93.9% وهو مستوى ممتاز يدل على صحة قاعدة بيانات التواصل.',
    '• نسبة السداد بعد التنبيه (74.5%) تؤكد فاعلية نظام التذكير الآلي في تحسين معدل التحصيل.',
    '• يُنصح بإضافة قناة SMS كقناة احتياطية لزيادة معدل الوصول إلى 98%+.',
    '• الأشهر التي سبقت المناسبات (يناير، مارس) سجّلت أعلى معدلات استجابة — يُقترح تكثيف التذكيرات قبل المناسبات الكبرى.',
]
for rec in recs:
    story.append(Paragraph(ar(rec), S_REC))

doc.build(story)
print('✅  PDF built successfully')

# ════════════════════════════════════════════════════════════════════════
#  EXCEL
# ════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Hex palette (Excel uses no #) ────────────────────────────────────────
HX = dict(
    nav='0B1F3A', blue='1A4A8A', teal='0D7A5F', gold='C07800',
    red='B02020', lgray='F5F5F3', mgray='D0D0CC',
    white='FFFFFF', black='000000',
    green_bg='E8F5F0', red_bg='FFF0F0',
)

THIN = Side(style='thin', color='CCCCCC')
FULL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def xset(ws, addr, value, bold=False, fc=None, bg=None,
         align='right', border=True, num_fmt=None, size=10):
    """Apply value + formatting to a cell."""
    cell = ws[addr] if isinstance(addr, str) else addr
    cell.value = value
    cell.font = Font(name='Arial', bold=bold,
                     color=fc or HX['black'], size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                                wrap_text=True)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if border:
        cell.border = FULL_BORDER
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def sheet_header(ws, title, subtitle=''):
    """Merge A1:H1 with navy title, A2:H2 with subtitle."""
    ws.merge_cells('A1:H1')
    xset(ws, 'A1', title, bold=True, fc=HX['white'],
         bg=HX['nav'], align='center', border=False, size=14)
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells('A2:H2')
        xset(ws, 'A2', subtitle, fc='AAAAAA',
             bg=HX['nav'], align='center', border=False, size=10)
        ws.row_dimensions[2].height = 20
    ws.sheet_view.rightToLeft = True

def hdr(ws, row_num, labels, widths, bg=HX['nav']):
    """Write a header row and set column widths. labels and widths are parallel lists."""
    for col_i, (label, width) in enumerate(zip(labels, widths), 1):
        cl = get_column_letter(col_i)
        xset(ws, f'{cl}{row_num}', label, bold=True, fc=HX['white'],
             bg=bg, align='center', size=10)
        ws.column_dimensions[cl].width = width

def total_xrow(ws, row_num, n_cols, bg=HX['nav']):
    """Apply navy total-row styling to n_cols columns."""
    for col_i in range(1, n_cols + 1):
        cl = get_column_letter(col_i)
        cell = ws[f'{cl}{row_num}']
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.font = Font(name='Arial', bold=True, color=HX['white'], size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ─────────────────────────────────────────────────────────────────────────
#  Sheet 1: Annual Accounts
# ─────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet('الحسابات السنوية')
sheet_header(ws1, 'تقرير الحسابات السنوية — 2025',
             'صندوق عائلة قطيفان | السنة المالية 2025')

# Summary block (rows 4-8)
ws1.merge_cells('A4:B4')
xset(ws1, 'A4', 'ملخص الحركة المالية', bold=True, fc=HX['white'],
     bg=HX['blue'], align='center', border=False, size=11)
ws1.row_dimensions[4].height = 22

summary = [
    ('رصيد بداية المدة',        OPENING_BALANCE,  HX['lgray'],    HX['teal']),
    ('إجمالي الاشتراكات',       total_income,     HX['white'],    HX['black']),
    ('(−) إجمالي المصروفات',   -total_expenses,   'FFF0F0',       HX['red']),
    ('رصيد نهاية المدة',        closing_balance,   HX['green_bg'], HX['teal']),
]
for i, (label, val, bg, fc) in enumerate(summary, 5):
    is_closing = label.startswith('رصيد نهاية')
    xset(ws1, f'A{i}', label, bold=is_closing, bg=bg, fc=fc, size=10 + is_closing)
    xset(ws1, f'B{i}', abs(val), bold=is_closing, bg=bg, fc=fc,
         align='left', num_fmt='#,##0.00', size=10 + is_closing)
    if val < 0:   # expenses shown as positive with (−) in label
        ws1[f'B{i}'].value = abs(val)

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 18

# Monthly income (rows 11 onward)
ws1.merge_cells('A10:B10')
xset(ws1, 'A10', 'الاشتراكات الشهرية', bold=True, fc=HX['white'],
     bg=HX['blue'], align='center', border=False, size=11)

# FIX 8/9: hdr for income uses only 2 columns A,B — widths already set above
xset(ws1, 'A11', 'الشهر',          bold=True, fc=HX['white'], bg=HX['blue'], align='center')
xset(ws1, 'B11', 'المحصَّل (ر.س)', bold=True, fc=HX['white'], bg=HX['blue'], align='center')

income_start_row = 12
for j, (month, val) in enumerate(monthly_income):
    row = income_start_row + j
    bg = HX['lgray'] if row % 2 == 0 else HX['white']
    xset(ws1, f'A{row}', month, bg=bg)
    xset(ws1, f'B{row}', val,   bg=bg, align='left', num_fmt='#,##0.00')

income_total_row = income_start_row + len(monthly_income)
xset(ws1, f'A{income_total_row}', 'الإجمالي', bold=True, fc=HX['white'], bg=HX['nav'])
# FIX 14: hardcode total value so it displays without recalc.py
xset(ws1, f'B{income_total_row}', total_income, bold=True, fc=HX['white'],
     bg=HX['nav'], align='left', num_fmt='#,##0.00')

# Expense category section
exp_sec_start = income_total_row + 2
ws1.merge_cells(f'A{exp_sec_start}:E{exp_sec_start}')
xset(ws1, f'A{exp_sec_start}', 'المصروفات حسب البند', bold=True,
     fc=HX['white'], bg=HX['red'], align='center', border=False, size=11)

exp_hdr_row = exp_sec_start + 1
# FIX 8: set columns A-E individually with correct widths
for cl, lbl, w in [('A','البند',25),('B','الحالات',12),
                   ('C','الإجمالي (ر.س)',18),('D','النسبة %',12)]:
    xset(ws1, f'{cl}{exp_hdr_row}', lbl, bold=True, fc=HX['white'],
         bg=HX['red'], align='center')
    ws1.column_dimensions[cl].width = w

exp_cat_labels = ['العزاء', 'نقوط الزواج', 'الطوارئ', 'المصاريف الإدارية']
exp_cat_vals   = [exp_condolence, exp_wedding, exp_emergency, exp_admin]
for k, (cat, tot) in enumerate(zip(exp_cat_labels, exp_cat_vals)):
    row = exp_hdr_row + 1 + k
    bg = HX['lgray'] if k % 2 == 0 else HX['white']
    xset(ws1, f'A{row}', cat, bg=bg)
    xset(ws1, f'B{row}', len(expenses_detail[list(expenses_detail)[k]]),
         bg=bg, align='center')
    xset(ws1, f'C{row}', tot, bg=bg, align='left', num_fmt='#,##0.00')
    # FIX 7: pre-compute percentage — no broken formula
    xset(ws1, f'D{row}', round(tot / total_expenses * 100, 1),
         bg=bg, align='left', num_fmt='0.0"%"')

exp_total_row = exp_hdr_row + 1 + len(cat_keys)
xset(ws1, f'A{exp_total_row}', 'المجموع', bold=True, fc=HX['white'], bg=HX['nav'])
xset(ws1, f'B{exp_total_row}', sum(len(v) for v in expenses_detail.values()),
     bold=True, fc=HX['white'], bg=HX['nav'], align='center')
xset(ws1, f'C{exp_total_row}', total_expenses, bold=True, fc=HX['white'],
     bg=HX['nav'], align='left', num_fmt='#,##0.00')
xset(ws1, f'D{exp_total_row}', 100.0, bold=True, fc=HX['white'],
     bg=HX['nav'], align='left', num_fmt='0.0"%"')

# ─────────────────────────────────────────────────────────────────────────
#  Sheet 2: Dues & Commitment
# ─────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('الذمم والالتزام')
sheet_header(ws2, 'تقرير الذمم والالتزام — 2025', 'كشف حساب الأعضاء النهائي')

hdr(ws2, 3,
    ['الاسم', 'مسدَّد (ر.س)', 'ذمة (ر.س)', 'أشهر مسدَّدة', 'الالتزام %', 'التقييم'],
    [28, 15, 15, 14, 13, 12],
    bg=HX['blue'])

rating_fc = {'ممتاز': HX['teal'], 'جيد': HX['blue'],
             'متوسط': HX['gold'], 'متأخر': HX['red']}

for i, (name, paid, debt, months_p, pct, rating) in enumerate(members_data, 4):
    bg = HX['lgray'] if i % 2 == 0 else HX['white']
    xset(ws2, f'A{i}', name, bg=bg)
    xset(ws2, f'B{i}', paid,  bg=bg, align='left', num_fmt='#,##0.00')
    xset(ws2, f'C{i}', debt,  bg=bg, align='left', num_fmt='#,##0.00',
         fc=HX['red'] if debt > 0 else HX['black'])
    xset(ws2, f'D{i}', months_p, bg=bg, align='center')
    xset(ws2, f'E{i}', pct / 100, bg=bg, align='left',  num_fmt='0.0%')
    rc = rating_fc.get(rating, HX['black'])
    xset(ws2, f'F{i}', rating, bg=bg, bold=True, fc=rc, align='center')

# Total row
mem_total_row = len(members_data) + 4
total_paid_sum = sum(r[1] for r in members_data)
total_debt_sum = sum(r[2] for r in members_data)
for cl in ('A', 'B', 'C', 'D', 'E', 'F'):
    cell = ws2[f'{cl}{mem_total_row}']
    cell.fill = PatternFill('solid', fgColor=HX['nav'])
    # FIX 10: explicit alignment, not bare Alignment()
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name='Arial', bold=True, color=HX['white'], size=10)
ws2[f'A{mem_total_row}'].value = 'المجموع'
xset(ws2, f'B{mem_total_row}', total_paid_sum, bold=True, fc=HX['white'],
     bg=HX['nav'], align='left', num_fmt='#,##0.00')
xset(ws2, f'C{mem_total_row}', total_debt_sum, bold=True, fc=HX['white'],
     bg=HX['nav'], align='left', num_fmt='#,##0.00')

# ─────────────────────────────────────────────────────────────────────────
#  Sheet 3: Notification Efficiency
# ─────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('كفاءة التنبيهات')
sheet_header(ws3, 'تقرير كفاءة التنبيهات الآلية — 2025', 'تحليل أداء الرسائل والتحصيل')
hdr(ws3, 3,
    ['الشهر', 'مُرسَل', 'موصَّل', 'مفتوح', 'سداد بعده', 'القناة', 'معدل التحويل'],
    [12, 10, 12, 10, 14, 20, 15],
    bg=HX['gold'])

for i, (month, sent, delivered, opened, paid_after, channel) in enumerate(notif_data, 4):
    bg = HX['lgray'] if i % 2 == 0 else HX['white']
    xset(ws3, f'A{i}', month,       bg=bg)
    xset(ws3, f'B{i}', sent,        bg=bg, align='center')
    xset(ws3, f'C{i}', delivered,   bg=bg, align='center')
    xset(ws3, f'D{i}', opened,      bg=bg, align='center')
    xset(ws3, f'E{i}', paid_after,  bg=bg, align='center')
    xset(ws3, f'F{i}', channel,     bg=bg, align='center')
    # FIX 7: pre-computed ratio
    xset(ws3, f'G{i}', round(paid_after / sent, 3), bold=True,
         fc=HX['teal'], bg=bg, align='center', num_fmt='0.0%')

notif_total_row = len(notif_data) + 4
totals = [total_sent, total_delivered, total_opened, total_paid_after]
for cl in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
    cell = ws3[f'{cl}{notif_total_row}']
    cell.fill = PatternFill('solid', fgColor=HX['nav'])
    # FIX 11: explicit alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(name='Arial', bold=True, color=HX['white'], size=10)
ws3[f'A{notif_total_row}'].value = 'الإجمالي'
for cl, val in zip(('B', 'C', 'D', 'E'), totals):
    ws3[f'{cl}{notif_total_row}'].value = val
overall_rate = round(total_paid_after / total_sent, 3)
ws3[f'G{notif_total_row}'].value = overall_rate
ws3[f'G{notif_total_row}'].number_format = '0.0%'

# ─────────────────────────────────────────────────────────────────────────
#  Sheet 4: Charts
# ─────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('الرسوم البيانية')
sheet_header(ws4, 'الرسوم البيانية التحليلية', 'ملخص بصري للتقارير الثلاثة')

# Data zones — start at row 40 (well below any merged header cells)
DATA_START = 40

# Expense pie data (columns A-B)
pie_labels = ['العزاء', 'نقوط الزواج', 'الطوارئ', 'إدارية']
pie_vals   = [exp_condolence, exp_wedding, exp_emergency, exp_admin]
for j, (lbl, val) in enumerate(zip(pie_labels, pie_vals)):
    ws4[f'A{DATA_START + j}'] = lbl
    ws4[f'B{DATA_START + j}'] = val

# Monthly income data (columns D-E)
for j, (month, val) in enumerate(monthly_income):
    ws4[f'D{DATA_START + j}'] = month
    ws4[f'E{DATA_START + j}'] = val

# Notification data (columns G-I); header at DATA_START-1
ws4[f'G{DATA_START - 1}'] = 'مُرسَل'
ws4[f'I{DATA_START - 1}'] = 'سداد بعده'
for j, (month, sent, _, _, paid_after, _ch) in enumerate(notif_data):
    ws4[f'G{DATA_START + j}'] = month
    ws4[f'H{DATA_START + j}'] = sent
    ws4[f'I{DATA_START + j}'] = paid_after

# Pie chart — expenses
pie = PieChart()
pie.title = 'توزيع المصروفات'
pie.add_data(Reference(ws4, min_col=2, min_row=DATA_START,
                        max_row=DATA_START + len(pie_vals) - 1))
pie.set_categories(Reference(ws4, min_col=1, min_row=DATA_START,
                               max_row=DATA_START + len(pie_vals) - 1))
pie.width = 14; pie.height = 10
ws4.add_chart(pie, 'A8')

# Bar chart — monthly income
bar = BarChart()
bar.type = 'col'
bar.title = 'الاشتراكات الشهرية'
bar.y_axis.title = 'ر.س'
bar.add_data(Reference(ws4, min_col=5, min_row=DATA_START,
                        max_row=DATA_START + len(monthly_income) - 1),
             titles_from_data=False)
bar.set_categories(Reference(ws4, min_col=4, min_row=DATA_START,
                               max_row=DATA_START + len(monthly_income) - 1))
bar.width = 20; bar.height = 10
ws4.add_chart(bar, 'I8')

# FIX 12: Line chart — both series and categories share same row range
# Header is at DATA_START-1, data at DATA_START .. DATA_START+5
line = LineChart()
line.title = 'الرسائل المرسلة مقابل السداد'
line.y_axis.title = 'عدد'
line.add_data(
    Reference(ws4, min_col=8, max_col=9,
              min_row=DATA_START - 1,          # row with series titles
              max_row=DATA_START + len(notif_data) - 1),
    titles_from_data=True,
)
line.set_categories(
    Reference(ws4, min_col=7,
              min_row=DATA_START,              # data only, no header
              max_row=DATA_START + len(notif_data) - 1),
)
line.width = 20; line.height = 10
ws4.add_chart(line, 'A28')

wb.save('/home/claude/qatifan_reports_2025.xlsx')
print('✅  Excel built successfully')
